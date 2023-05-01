import os
import io
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox

from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.layout import LAParams, LTTextBox, LTTextLine, LTFigure, LTImage, LTChar
from pdfminer.converter import PDFPageAggregator
from openpyxl import Workbook
from PIL import Image
import pytesseract

g_text = None


def extract_image_from_pdf(file_name):
    # delete existed .jpg
    for path in os.listdir(os.getcwd()):
        if ".xlsx" in path:
            os.remove(path)

    # Set parameters for analysis.
    laparams = LAParams()
    rsrcmgr = PDFResourceManager()

    # Create a PDF page aggregator object.
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)

    # Open a PDF file.
    fp = open(file_name, 'rb')

    # Create a PDF parser object associated with the file object.
    parser = PDFParser(fp)
    # Create a PDF document object that stores the document structure.
    # Supply the password for initialization.
    document = PDFDocument(parser)

    for p, page in enumerate(PDFPage.create_pages(document)):
        interpreter.process_page(page)
        # receive the LTPage object for the page.
        layout = device.get_result()

        for node in layout:
            # print(type(node))

            if isinstance(node, LTFigure):
                for n in node:
                    # print(" {}".format(type(n)))
                    if n.stream:
                        buffer = io.BytesIO(n.stream.get_data())
                        # with open("test_data_{}.img".format(p),'wb') as fp :
                        #    fp.write(buffer.read())
                        pillow_object = Image.open(buffer)
                        pillow_object.save("test_image_{}.jpg".format(p))


def handle_invoice(text, sheet):
    for temp in text.split('\n'):
        row = temp.split()
        sheet.append(row)


g_start = False
g_max_num = 0


def handle_packaging_list(text, sheet):
    row = None
    global g_start
    global g_max_num
    for temp in text.split('\n'):
        if ("CARTON" in temp) or ("PALLET" in temp) and (not g_start):
            g_start = True
            elements = temp.split()
            g_max_num = len(elements)
            # print(temp)
            # print("g_max_num: {0}".format(g_max_num))

        # handle   @xx @xx @xx
        if g_start and (len(temp.split()) < g_max_num - 2) and ("@" in temp):
            row = temp.split()
            row.insert(0, None)
            row.insert(0, None)
            row.insert(0, None)

        # handle xx space space xx xx
        elif g_start and (len(temp.split()) < g_max_num - 2) and ("@" not in temp):
            row = temp.split()
            row.insert(0, None)
            row.insert(0, None)
            row.insert(0, None)
            row.insert(4, None)
            row.insert(4, None)
        else:
            # handle normal line
            row = temp.split()

        # print(row)
        sheet.append(row)


def handle_other_work(text, sheet):
    print("Currently, don't support such kind of pdf format.")


def extract_text_and_and_write_to_excel(file_name):
    # extract text
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    work_count = 0
    works = ["invoice", "packing_list"]

    # delete existed .xlsx
    for path in os.listdir(os.getcwd()):
        if ".xlsx" in path:
            os.remove(path)

    work_book = Workbook()
    process_work = None
    sheet = None

    for path in os.listdir(os.getcwd()):
        if ".jpg" in path:
            file = Image.open(path)
            config = r'-c tessedit_char_blacklist=  --psm 6'
            content = pytesseract.image_to_string(file, config=config)
            # print(context)

            if "SHENGGAO" in content:
                print("work_count:{0}".format(work_count))
                process_work = works[work_count]
                work_count += 1

            # create sheet
            if process_work not in work_book.sheetnames:
                sheet = work_book.create_sheet(process_work)

            # process image context
            match process_work:
                case "invoice":
                    handle_invoice(content, sheet)

                case "packing_list":
                    handle_packaging_list(content, sheet)

                case _:
                    handle_other_work(content, sheet)

    print("Finish handling packing list pages.")
    del work_book["Sheet"]
    work_book.save("{0}.xlsx".format(file_name.split(".")[0]))
    print("All the data has been save to {0}.xlsx".format(file_name.split(".")[0]))


def open_file():
    try:
        # open file and return full path
        path = filedialog.askopenfilename()
        file_path.set(path)
    except Exception as e:
        print("You haven't selected any file yet", e)
        file_path.set("")
    print("filename is {0}".format(file_path.get()))
    text.delete(1.0, END)
    start_button.config(bg='SystemButtonFace')


def extract_pdf():
    if file_path.get() == "" or ".pdf" not in file_path.get():
        messagebox.showerror("Error", "Please select PDF file")
    else:
        (directory, file_name) = os.path.split(file_path.get())
        print("directory:{0}, file_name:{1}, file_path:{2}".format(directory, file_name, file_path.get()))
        extract_image_from_pdf(file_path.get())
        extract_text_and_and_write_to_excel(file_name)
        info = "Finish converting PDF to Excel. Excel is stored at {0}".format(directory+"/"+file_name.split(".")[0]
                                                                               +".xlxs")
        text.delete(1.0, END)
        text.insert(INSERT, info)
    start_button.config(bg='green')


if __name__ == "__main__":
    root_window = Tk()
    root_window.geometry('450x300')
    root_window.title("ExtractPdf")
    file_path = tk.StringVar()

    entry = tk.Entry(root_window, width=40, textvariable=file_path)
    entry.grid(row=0, column=1, padx=5, pady=5)

    text = tk.Text(root_window, name="text", width=40, height=10, padx=5, pady=5, undo=True, autoseparators=False)
    text.grid(row=2, column=1)

    browse_button = tk.Button(root_window, text='BrowserFile', padx=5, pady=5, command=open_file)
    browse_button.grid(row=0, column=2)

    start_button = tk.Button(root_window, width=40, text='Start', padx=5, pady=5, command=extract_pdf)
    start_button.grid(row=1, columnspan=2)

    root_window.mainloop()
